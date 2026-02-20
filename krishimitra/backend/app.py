from flask import Flask, request, jsonify
from flask_sqlalchemy import SQLAlchemy
from flask_bcrypt import Bcrypt
from flask_cors import CORS
from flask_jwt_extended import JWTManager, create_access_token, jwt_required, get_jwt_identity
import os
from datetime import datetime
import uuid
from dotenv import load_dotenv
import random
import json
import csv
from io import BytesIO, StringIO
import requests
from openpyxl import load_workbook
import base64
import mimetypes

app = Flask(__name__)
CORS(app)
load_dotenv()

# Configuration
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'your-secret-key-change-this-in-production')
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URL', 'sqlite:///krishimitra.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['JWT_SECRET_KEY'] = os.getenv('JWT_SECRET_KEY', 'jwt-secret-key-change-this')

# Initialize extensions
db = SQLAlchemy(app)
bcrypt = Bcrypt(app)
jwt = JWTManager(app)

# Create upload folder if it doesn't exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Models
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    phone = db.Column(db.String(20), nullable=True)
    trust_score = db.Column(db.Integer, default=0)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    def __repr__(self):
        return f'<User {self.username}>'

class UploadedFile(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    filename = db.Column(db.String(255), nullable=False)
    original_filename = db.Column(db.String(255), nullable=False)
    file_path = db.Column(db.String(500), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    uploaded_at = db.Column(db.DateTime, default=datetime.utcnow)
    processed = db.Column(db.Boolean, default=False)
    ai_response = db.Column(db.Text, nullable=True)
    
    user = db.relationship('User', backref=db.backref('files', lazy=True))

class SensorReport(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    filename = db.Column(db.String(255), nullable=False)
    original_filename = db.Column(db.String(255), nullable=False)
    file_path = db.Column(db.String(500), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    trust_score_10 = db.Column(db.Float, default=0.0)
    address_text = db.Column(db.String(500), nullable=True)
    lat = db.Column(db.Float, nullable=True)
    lon = db.Column(db.Float, nullable=True)
    ai_summary = db.Column(db.Text, nullable=True)

# Routes
@app.route('/')
def home():
    return jsonify({'message': 'Krishimitra Backend API'})

@app.route('/api/register', methods=['POST'])
def register():
    data = request.get_json() or {}
    email = data.get('email', '').strip()
    password = data.get('password', '')
    name = (data.get('name') or data.get('username') or '').strip()
    phone = (data.get('phone') or '').strip()
    trust_score = data.get('trustScore')
    if not email or not password:
        return jsonify({'error': 'Email and password are required'}), 400
    if User.query.filter_by(email=email).first():
        return jsonify({'error': 'Email already exists'}), 400
    hashed_password = bcrypt.generate_password_hash(password).decode('utf-8')
    if trust_score is None:
        trust_score = random.randint(60, 100)
    new_user = User(
        username=name or email.split('@')[0],
        email=email,
        password=hashed_password,
        phone=phone,
        trust_score=trust_score
    )
    db.session.add(new_user)
    db.session.commit()
    return jsonify({
        'message': 'User created successfully',
        'user': {
            'id': new_user.id,
            'name': new_user.username,
            'email': new_user.email,
            'phone': new_user.phone,
            'trustScore': new_user.trust_score
        }
    }), 201

@app.route('/api/login', methods=['POST'])
def login():
    data = request.get_json() or {}
    email = data.get('email', '').strip()
    password = data.get('password', '')
    if not email or not password:
        return jsonify({'error': 'Email and password are required'}), 400
    user = User.query.filter_by(email=email).first()
    if not user or not bcrypt.check_password_hash(user.password, password):
        return jsonify({'error': 'Invalid email or password'}), 401
    access_token = create_access_token(identity=user.id)
    return jsonify({
        'message': 'Login successful',
        'access_token': access_token,
        'user': {
            'id': user.id,
            'name': user.username,
            'email': user.email,
            'phone': user.phone,
            'trustScore': user.trust_score
        }
    }), 200

def _mime_for_ext(ext):
    m = mimetypes.types_map.get(ext.lower()) if ext else None
    if not m:
        return 'application/octet-stream'
    return m

def _analyze_with_openai(path, prompt, crop=None):
    key = os.getenv('OPENAI_API_KEY')
    if not key:
        return None
    try:
        with open(path, 'rb') as f:
            b64 = base64.b64encode(f.read()).decode('ascii')
        ext = os.path.splitext(path)[1]
        mime = _mime_for_ext(ext)
        data_uri = f"data:{mime};base64,{b64}"
        disease_map = {
            'wheat': ['rust (stripe, leaf, stem)', 'powdery mildew', 'leaf blight', 'aphids'],
            'rice': ['rice blast', 'bacterial leaf blight', 'sheath blight', 'brown planthopper'],
            'maize': ['northern leaf blight', 'common rust', 'gray leaf spot', 'fall armyworm', 'common smut'],
        }
        crop_key = (crop or '').strip().lower()
        targets = disease_map.get(crop_key, [])
        target_text = ", ".join(targets) if targets else "common crop diseases and pests"
        system = (
            "You are an agronomy assistant. Analyze the crop photo for health, diseases, pests, and damage. "
            f"Focus on {crop or 'the crop'} and especially: {target_text}. "
            "Be highly sensitive to small signs: brown patches, irregular patterns, holes, spots, edge burn, leaf curling, chlorosis, necrosis, webbing, and insect bite marks. "
            "Return strict JSON with keys: "
            "summary (string), details (string), confidence (Low|Medium|High), "
            "issues (array of objects: name, likelihood (0-100), description), "
            "observations (array of objects: type, description, severity (0-100), confidence (Low|Medium|High)), "
            "recommendations (array of strings)."
        )
        body = {
            'model': os.getenv('OPENAI_MODEL', 'gpt-4o-mini'),
            'messages': [
                {'role': 'system', 'content': system},
                {
                    'role': 'user',
                    'content': [
                        {'type': 'text', 'text': prompt or 'Assess crop condition and identify any disease or pest.'},
                        {'type': 'image_url', 'image_url': {'url': data_uri}},
                    ],
                },
            ],
            'response_format': {'type': 'json_object'},
        }
        r = requests.post('https://api.openai.com/v1/chat/completions', headers={
            'Authorization': f'Bearer {key}',
            'Content-Type': 'application/json',
        }, json=body, timeout=60)
        j = r.json()
        content = j.get('choices', [{}])[0].get('message', {}).get('content')
        if not content:
            return None
        parsed = json.loads(content)
        return parsed
    except Exception:
        return None

@app.route('/api/crop-analysis', methods=['POST'])
def crop_analysis():
    if 'image' not in request.files:
        return jsonify({'error': 'No image provided'}), 400
    image = request.files['image']
    if image.filename == '':
        return jsonify({'error': 'No image selected'}), 400
    ext = os.path.splitext(image.filename)[1]
    unique = f"{uuid.uuid4().hex}{ext}"
    path = os.path.join(app.config['UPLOAD_FOLDER'], unique)
    image.save(path)
    data = request.form.to_dict() if request.form else {}
    prompt = data.get('prompt')
    crop = data.get('crop')
    ai = _analyze_with_openai(path, prompt, crop)
    if ai and isinstance(ai, dict):
        return jsonify({
            'summary': ai.get('summary') or 'Analysis available',
            'details': ai.get('details') or '',
            'confidence': ai.get('confidence') or 'Medium',
            'issues': ai.get('issues') or [],
            'observations': ai.get('observations') or [],
            'recommendations': ai.get('recommendations') or [],
        }), 200
    return jsonify({
        'summary': 'Likely healthy',
        'details': 'Leaves appear normal. No obvious signs of damage or disease detected in this demo response.',
        'confidence': 'Medium',
        'issues': [],
        'observations': [],
        'recommendations': [],
    }), 200

@app.route('/api/stage-verify', methods=['POST'])
def stage_verify():
    if 'image' not in request.files:
        return jsonify({'error': 'No image provided'}), 400
    image = request.files['image']
    if image.filename == '':
        return jsonify({'error': 'No image selected'}), 400
    stage = (request.form.get('stage') or '').strip()
    crop = (request.form.get('crop') or '').strip()
    ext = os.path.splitext(image.filename)[1]
    unique = f"{uuid.uuid4().hex}{ext}"
    path = os.path.join(app.config['UPLOAD_FOLDER'], unique)
    image.save(path)
    prompts = {
        '1': 'Verify this is seeds/sowing stage (seed packets, seeds in soil, early seedlings).',
        '2': 'Verify this is growth stage (green crops, leaves, stems, field with growing plants).',
        '3': 'Verify this is harvest stage (harvested crop, grain, bundles, transport).',
    }
    pfx = prompts.get(stage, 'Assess crop stage (seed, growth, harvest).')
    ai = _analyze_with_openai(path, pfx, crop)
    awarded = 0
    reason = 'No points'
    if ai and isinstance(ai, dict):
        conf = ai.get('confidence')
        issues = ai.get('issues') or []
        observations = ai.get('observations') or []
        text = json.dumps({'issues': issues, 'observations': observations}).lower()
        ok = False
        if stage == '1':
            ok = ('seed' in text) or ('sowing' in text) or ('seedling' in text)
        elif stage == '2':
            ok = ('leaf' in text) or ('growth' in text) or ('stem' in text) or ('green' in text)
        elif stage == '3':
            ok = ('harvest' in text) or ('grain' in text) or ('bundle' in text) or ('transport' in text)
        if ok:
            if (conf or '').lower() == 'high':
                awarded = 10
                reason = 'Stage verified with high confidence'
            else:
                awarded = 8
                reason = 'Stage verified'
        else:
            awarded = 5
            reason = 'Stage unclear, partial points'
        return jsonify({
            'stage': stage,
            'awardedPoints': awarded,
            'reason': reason,
            'analysis': {
                'summary': ai.get('summary'),
                'details': ai.get('details'),
                'confidence': ai.get('confidence'),
                'issues': ai.get('issues'),
                'observations': ai.get('observations'),
                'recommendations': ai.get('recommendations'),
            }
        }), 200
    return jsonify({
        'stage': stage,
        'awardedPoints': 10,
        'reason': 'Demo award',
        'analysis': {
            'summary': 'Likely healthy',
            'details': 'Demo response.',
            'confidence': 'Medium',
            'issues': [],
            'observations': [],
            'recommendations': [],
        }
    }), 200

def _count_small_transactions_csv(path, threshold=500.0):
    count = 0
    total = 0.0
    lines = 0
    try:
        with open(path, 'r', encoding='utf-8', errors='ignore') as f:
            reader = csv.DictReader(f)
            for row in reader:
                amount = row.get('amount') or row.get('Amount') or row.get('AMOUNT')
                if amount is None:
                    continue
                try:
                    val = float(str(amount).replace(',', '').strip())
                except Exception:
                    continue
                total += abs(val)
                lines += 1
                if abs(val) <= threshold:
                    count += 1
    except Exception:
        pass
    return count, total, lines

def _count_small_transactions_xlsx(path, threshold=500.0):
    count = 0
    total = 0.0
    lines = 0
    try:
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        headers = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(h).lower() if h is not None else '' for h in row]
                continue
            # find amount-like column
            idx = None
            for j, h in enumerate(headers):
                if 'amount' in h or 'amt' in h or h in ('debit', 'credit'):
                    idx = j
                    break
            if idx is None:
                continue
            val = row[idx]
            try:
                num = float(str(val).replace(',', '').strip())
            except Exception:
                continue
            total += abs(num)
            lines += 1
            if abs(num) <= threshold:
                count += 1
    except Exception:
        pass
    return count, total, lines

def _count_small_transactions_json(path, threshold=500.0):
    count = 0
    total = 0.0
    lines = 0
    try:
        data = json.loads(open(path, 'r', encoding='utf-8').read())
        def walk(x):
            nonlocal count, total, lines
            if isinstance(x, dict):
                for k, v in x.items():
                    if k.lower() in ('amount', 'amt', 'debit', 'credit') and isinstance(v, (int, float, str)):
                        try:
                            num = float(str(v).replace(',', '').strip())
                        except Exception:
                            continue
                        total += abs(num)
                        lines += 1
                        if abs(num) <= threshold:
                            count += 1
                    else:
                        walk(v)
            elif isinstance(x, list):
                for v in x:
                    walk(v)
        walk(data)
    except Exception:
        pass
    return count, total, lines

def _count_small_transactions_pdf(path, threshold=500.0):
    count = 0
    total = 0.0
    lines = 0
    try:
        # Best-effort text extraction without external libs: decode bytes and regex amounts
        raw = open(path, 'rb').read()
        try:
            text = raw.decode('utf-8', errors='ignore')
        except Exception:
            text = raw.decode('latin-1', errors='ignore')
        import re
        # Match currency-like numbers: optional ₹/Rs/INR, commas, decimals, optional sign
        pattern = re.compile(r"(?:₹\s*|Rs\.?\s*|INR\s*)?(-?\d{1,3}(?:,\d{3})*(?:\.\d+)?|-?\d+(?:\.\d+)?)")
        matches = pattern.findall(text)
        for m in matches:
            try:
                num = float(str(m).replace(',', '').strip())
            except Exception:
                continue
            total += abs(num)
            lines += 1
            if abs(num) <= threshold:
                count += 1
    except Exception:
        pass
    return count, total, lines

@app.route('/api/bank-statement', methods=['POST'])
def bank_statement():
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    f = request.files['file']
    if f.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    ext = os.path.splitext(f.filename)[1].lower()
    unique = f"{uuid.uuid4().hex}{ext}"
    path = os.path.join(app.config['UPLOAD_FOLDER'], unique)
    f.save(path)
    small = 0
    total = 0.0
    lines = 0
    if ext in ('.csv',):
        small, total, lines = _count_small_transactions_csv(path)
    elif ext in ('.xlsx', '.xlsm', '.xltx', '.xltm'):
        small, total, lines = _count_small_transactions_xlsx(path)
    elif ext in ('.json', '.txt'):
        small, total, lines = _count_small_transactions_json(path)
    elif ext in ('.pdf',):
        small, total, lines = _count_small_transactions_pdf(path)
    else:
        return jsonify({'error': 'Unsupported file type'}), 400
    ratio = (small / lines) if lines else 0.0
    active = (small >= 15) or (ratio >= 0.5)
    delta = 20 if active else 0
    return jsonify({
        'active': active,
        'smallTransactions': small,
        'totalTransactions': lines,
        'activityRatio': round(ratio, 2),
        'trustDelta': delta,
    }), 200

def _normalize_metric(key, value):
    try:
        v = float(value)
    except Exception:
        return None
    k = (key or '').lower()
    if 'soil' in k and 'moist' in k:
        return max(0.0, min(10.0, v / 10.0))
    if k in ('ph',) or 'ph' in k:
        return max(0.0, min(10.0, 10.0 - abs(v - 6.5) * (10.0 / 6.5)))
    if 'humid' in k:
        return max(0.0, min(10.0, v / 10.0))
    if 'temp' in k:
        return max(0.0, min(10.0, (50.0 - abs(v - 25.0)) / 5.0))
    if 'rain' in k or 'precip' in k:
        return max(0.0, min(10.0, 10.0 - max(0.0, v - 20.0) * 0.5))
    if 'wind' in k:
        return max(0.0, min(10.0, 10.0 - v * 0.5))
    if 0.0 <= v <= 100.0:
        return v / 10.0
    return max(0.0, min(10.0, v / 10.0))

def _extract_address_and_coords(data):
    addr = None
    lat = None
    lon = None
    def _try_get(obj, keys):
        for k in keys:
            if isinstance(obj, dict) and k in obj and obj[k]:
                return obj[k]
        return None
    if isinstance(data, dict):
        addr = _try_get(data, ['address', 'location', 'field_address'])
        lat = _try_get(data, ['lat', 'latitude'])
        lon = _try_get(data, ['lon', 'lng', 'longitude'])
    if isinstance(lat, str):
        try:
            lat = float(lat)
        except Exception:
            lat = None
    if isinstance(lon, str):
        try:
            lon = float(lon)
        except Exception:
            lon = None
    if (not lat or not lon) and addr:
        try:
            r = requests.get('https://nominatim.openstreetmap.org/search', params={'q': addr, 'format': 'json', 'limit': 1}, headers={'User-Agent': 'krishimitra-app'})
            j = r.json()
            if isinstance(j, list) and j:
                lat = float(j[0]['lat'])
                lon = float(j[0]['lon'])
        except Exception:
            pass
    return addr, lat, lon

def _flatten_numeric(obj):
    nums = []
    metrics = {'ph': None, 'moisture': None, 'nitrogen': None}
    if isinstance(obj, dict):
        for k, v in obj.items():
            if isinstance(v, (int, float)):
                n = _normalize_metric(k, v)
                if n is not None:
                    nums.append(n)
            elif isinstance(v, (dict, list)):
                nums.extend(_flatten_numeric(v))
    elif isinstance(obj, list):
        for v in obj:
            if isinstance(v, (int, float)):
                n = _normalize_metric('', v)
                if n is not None:
                    nums.append(n)
            elif isinstance(v, (dict, list)):
                nums.extend(_flatten_numeric(v))
    return nums

@app.route('/api/sensor-readings', methods=['POST'])
def sensor_readings():
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    f = request.files['file']
    if f.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    ext = os.path.splitext(f.filename)[1].lower()
    unique = f"{uuid.uuid4().hex}{ext}"
    path = os.path.join(app.config['UPLOAD_FOLDER'], unique)
    f.save(path)
    nums = []
    metrics = {'ph': None, 'moisture': None, 'nitrogen': None}
    addr = None
    lat = None
    lon = None
    summary = None
    try:
        if ext in ('.json', '.txt'):
            data = json.loads(open(path, 'r', encoding='utf-8').read())
            nums = _flatten_numeric(data)
            addr, lat, lon = _extract_address_and_coords(data)
            def walk(obj):
                if isinstance(obj, dict):
                    for k, v in obj.items():
                        lk = str(k).lower()
                        if isinstance(v, (int, float)):
                            if metrics['ph'] is None and ('ph' in lk):
                                metrics['ph'] = float(v)
                            if metrics['moisture'] is None and ('moist' in lk or 'humidity' in lk):
                                metrics['moisture'] = float(v)
                            if metrics['nitrogen'] is None and ('nitrogen' in lk or 'n' == lk):
                                metrics['nitrogen'] = float(v)
                        elif isinstance(v, (dict, list)):
                            walk(v)
                elif isinstance(obj, list):
                    for v in obj:
                        walk(v)
            walk(data)
        elif ext in ('.csv',):
            content = open(path, 'r', encoding='utf-8', errors='ignore').read()
            reader = csv.DictReader(StringIO(content))
            for row in reader:
                for k, v in row.items():
                    try:
                        n = _normalize_metric(k, float(v))
                        if n is not None:
                            nums.append(n)
                        lk = str(k).lower()
                        fv = float(v)
                        if metrics['ph'] is None and ('ph' in lk):
                            metrics['ph'] = fv
                        if metrics['moisture'] is None and ('moist' in lk or 'humidity' in lk):
                            metrics['moisture'] = fv
                        if metrics['nitrogen'] is None and ('nitrogen' in lk or lk == 'n'):
                            metrics['nitrogen'] = fv
                    except Exception:
                        pass
            addr = None
        elif ext in ('.xlsx', '.xlsm', '.xltx', '.xltm'):
            wb = load_workbook(path, data_only=True)
            ws = wb.active
            headers = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(h) if h is not None else '' for h in row]
                    continue
                for k, v in zip(headers, row):
                    try:
                        n = _normalize_metric(k, float(v))
                        if n is not None:
                            nums.append(n)
                        lk = str(k).lower()
                        fv = float(v)
                        if metrics['ph'] is None and ('ph' in lk):
                            metrics['ph'] = fv
                        if metrics['moisture'] is None and ('moist' in lk or 'humidity' in lk):
                            metrics['moisture'] = fv
                        if metrics['nitrogen'] is None and ('nitrogen' in lk or lk == 'n'):
                            metrics['nitrogen'] = fv
                    except Exception:
                        pass
        else:
            return jsonify({'error': 'Unsupported file type'}), 400
    except Exception:
        pass
    trust = round(sum(nums) / max(1, len(nums)), 1) if nums else 5.0
    addr_text = None
    try:
        if addr is None:
            addr_text = None
        elif isinstance(addr, (dict, list)):
            addr_text = json.dumps(addr, ensure_ascii=False)
        else:
            addr_text = str(addr)
    except Exception:
        addr_text = None
    sr = SensorReport(
        filename=unique,
        original_filename=f.filename,
        file_path=path,
        trust_score_10=trust,
        address_text=addr_text,
        lat=lat,
        lon=lon,
        ai_summary=summary
    )
    db.session.add(sr)
    db.session.commit()
    return jsonify({
        'trustScore': trust,
        'address': addr,
        'lat': lat,
        'lon': lon,
        'reportId': sr.id,
        'metrics': metrics
    }), 200

@app.route('/api/weather', methods=['GET'])
def weather():
    address = request.args.get('address')
    lat = request.args.get('lat')
    lon = request.args.get('lon')
    if address and (not lat or not lon):
        try:
            r = requests.get('https://nominatim.openstreetmap.org/search', params={'q': address, 'format': 'json', 'limit': 1}, headers={'User-Agent': 'krishimitra-app'})
            j = r.json()
            if isinstance(j, list) and j:
                lat = j[0]['lat']
                lon = j[0]['lon']
        except Exception:
            pass
    if not lat or not lon:
        return jsonify({'error': 'lat/lon or address required'}), 400
    try:
        wx = requests.get('https://api.open-meteo.com/v1/forecast', params={
            'latitude': lat,
            'longitude': lon,
            'current_weather': 'true',
            'hourly': 'temperature_2m,precipitation,wind_speed_10m'
        })
        data = wx.json()
        current = data.get('current_weather') or {}
        hourly = data.get('hourly') or {}
        return jsonify({
            'current': current,
            'hourly': hourly,
            'lat': float(lat),
            'lon': float(lon)
        }), 200
    except Exception:
        return jsonify({'error': 'weather fetch failed'}), 500

@app.route('/api/upload', methods=['POST'])
@jwt_required()
def upload_file():
    current_user_id = get_jwt_identity()
    
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    # Generate unique filename
    file_extension = os.path.splitext(file.filename)[1]
    unique_filename = f"{uuid.uuid4().hex}{file_extension}"
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
    
    # Save file
    file.save(file_path)
    
    # Create file record in database
    uploaded_file = UploadedFile(
        filename=unique_filename,
        original_filename=file.filename,
        file_path=file_path,
        user_id=current_user_id
    )
    
    db.session.add(uploaded_file)
    db.session.commit()
    
    return jsonify({
        'message': 'File uploaded successfully',
        'file_id': uploaded_file.id,
        'filename': uploaded_file.original_filename
    }), 201

@app.route('/api/process/<int:file_id>', methods=['POST'])
@jwt_required()
def process_file(file_id):
    current_user_id = get_jwt_identity()
    
    uploaded_file = UploadedFile.query.filter_by(id=file_id, user_id=current_user_id).first()
    
    if not uploaded_file:
        return jsonify({'error': 'File not found'}), 404
    
    # Get processing prompt from request
    data = request.get_json()
    prompt = data.get('prompt', 'Analyze this agricultural data and provide insights.')
    
    # TODO: Integrate with AI API (OpenAI, Gemini, etc.)
    # This is where you'll call your AI service
    ai_response = f"Processed file: {uploaded_file.original_filename}\nPrompt: {prompt}\n\nAI Response: This is a placeholder response. Integrate with actual AI API."
    
    # Update file record
    uploaded_file.processed = True
    uploaded_file.ai_response = ai_response
    db.session.commit()
    
    return jsonify({
        'message': 'File processed successfully',
        'response': ai_response
    }), 200

@app.route('/api/files', methods=['GET'])
@jwt_required()
def get_user_files():
    current_user_id = get_jwt_identity()
    
    files = UploadedFile.query.filter_by(user_id=current_user_id).order_by(UploadedFile.uploaded_at.desc()).all()
    
    files_data = []
    for file in files:
        files_data.append({
            'id': file.id,
            'filename': file.original_filename,
            'uploaded_at': file.uploaded_at.isoformat(),
            'processed': file.processed,
            'ai_response': file.ai_response
        })
    
    return jsonify({'files': files_data}), 200

@app.route('/api/profile', methods=['GET'])
@jwt_required()
def get_profile():
    current_user_id = get_jwt_identity()
    
    user = User.query.get(current_user_id)
    
    if not user:
        return jsonify({'error': 'User not found'}), 404
    
    return jsonify({
        'id': user.id,
        'username': user.username,
        'email': user.email,
        'created_at': user.created_at.isoformat()
    }), 200

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True, port=5000)
